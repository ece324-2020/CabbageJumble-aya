"""
# Read an Excel file and get hyperparameters

From https://openpyxl.readthedocs.io/en/stable/pandas.html

Assumptions:
    - You are running it from scratch. It will read all of the hyperparameters
    - Manually set wb.Hyperparameters.Done = True when done (ideally automatic)

"""
from openpyxl import load_workbook
from itertools import islice
import pandas as pd



def read_hyperparameters(xlsx_path, worksheet: str = 'Hyperparameters'):
    """
    Read the Hyperparameters from a spreadsheet
    - Taken from https://openpyxl.readthedocs.io/en/stable/pandas.html

    :param xlsx_path: str - path to xlsx worksheet
    :param worksheet: str - name of Hyperparameter worksheet
    :return: pd.DataFrame - Pandas dataframe of info
    """

    wb = load_workbook(filename=xlsx_path)
    ws = wb[worksheet]
    wb.close()

    # From Documentation
    data = ws.values
    # First row is column names
    cols = next(data)[1:]
    # Extract Data Fields
    data = list(data)
    # Index is first column of data
    idx = [r[0] for r in data]
    data = (islice(r, 1, None) for r in data)
    df = pd.DataFrame(data, index=idx, columns=cols)

    return df


def write_results(results, xlsx_path, worksheet: str = 'Results'):
    """
    Write results
    :param results: dict - {'Trial': 1, 'Train Acc': 0.883, 'Train Loss': 0.546, 'Valid Acc':, 'Valid Loss':,
    'Test Acc':, 'Test Loss':}
    :param xlsx_path:
    :param worksheet:
    :return:
    """
    # Open and Read
    wb = load_workbook(filename=xlsx_path)
    ws = wb[worksheet]

    # Append Row
    trial = results.get('Trial', None)
    tr_acc = results.get('Train Acc', None)
    tr_loss = results.get('Train Loss', None)
    v_acc = results.get('Valid Acc', None)
    v_loss = results.get('Valid Loss', None)
    t_acc = results.get('Test Acc', None)
    t_loss = results.get('Test Loss', None)

    ws.append((trial, tr_acc, tr_loss, v_acc, v_loss, t_acc, t_loss))

    # Save and Close
    wb.save(filename=xlsx_path)
    wb.close()




if __name__ == '__main__':
    xlsx_path = '../$ scrap_data/test.xlsx'
    df = read_hyperparameters(xlsx_path)
    results = {'Trial': 1, 'Train Acc': 0.883, 'Train Loss': 0.546, 'Valid Acc': 0.797, 'Valid Loss': 0.721,
    'Test Acc': 0.657, 'Test Loss': 1.043}
    write_results(results, xlsx_path)



